import smtplib
from email.mime.text import MIMEText
import openpyxl
import datetime

def email_send(data):
    _name = data[0]
    _email = data[1]
    _subject = data[2]
    _content = data[3]

    _subject = _subject.replace("{이름}", _name)
    _content = _content.replace("{이름}", _name)

    try:
        s = smtplib.SMTP('smtp.naver.com', 587)
        s.starttls()
        s.login('naver_id@naver.com', 'app_password')
        msg = MIMEText(_content)
        msg['Subject'] = _subject
        msg['From'] = "naver_id@naver.com"
        msg['To'] = _email
        s.sendmail(msg['From'], msg['To'], msg.as_string())
        s.quit()
        return True

    except Exception as e:
        print(e)
        return False

wb = openpyxl.load_workbook("email_list.xlsx")
ws = wb.active
for row in ws.rows:
    rownum = row[0].row
    if rownum == 1:
        continue
    data = (row[0].value, row[1].value, row[2].value, row[3].value)

    if row[4].value == "성공":
        continue

    print("email sending --> ", row[0].value, row[1].value)
    result = email_send(data)

    dt = datetime.datetime.now()
    if result:
        ws['E'+str(rownum)] = "성공"
        ws['F'+str(rownum)] = dt.strftime("%Y-%m-%d %H:%M:%S")
    else:
        ws['E'+str(rownum)] = "실패"
        ws['F'+str(rownum)] = dt.strftime("%Y-%m-%d %H:%M:%S")

wb.save("email_list.xlsx")