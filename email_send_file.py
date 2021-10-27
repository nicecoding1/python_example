import os
import smtplib
from email.mime.text import MIMEText
import openpyxl
import datetime
from email.encoders import encode_base64
from email.header import Header
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.utils import formatdate

def email_send(data):
    _name = data[0]
    _email = data[1]
    _subject = data[2]
    _content = data[3]
    _file1 = data[4]
    _file2 = data[5]

    _subject = _subject.replace("{이름}", _name)
    _content = _content.replace("{이름}", _name)

    try:
        s = smtplib.SMTP('smtp.naver.com', 587)
        s.starttls()
        s.login('naver_id@naver.com', 'app_password')
        msg = MIMEMultipart()
        msg['Subject'] = Header(s=_subject, charset='utf-8')
        body = MIMEText(_content, _charset='utf-8')
        msg.attach(body)

        files = list()
        if _file1: files.append(_file1)
        if _file2: files.append(_file2)

        for f in files:
            part = MIMEBase('application', 'octect-stream')
            part.set_payload(open(f, "rb").read())
            encode_base64(part)
            part.add_header('Content-Disposition', 'attachment; filename="%s"' % os.path.basename(f))
            msg.attach(part)

        msg['From'] = "naver_id@naver.com"
        msg['To'] = _email
        msg['Date'] = formatdate(localtime=True)
        s.send_message(msg)
        s.quit()
        return True

    except Exception as e:
        print(e)
        return False

wb = openpyxl.load_workbook("email_list.xlsx")
ws = wb.active
for row in ws.rows:
    rownum = row[0].row
    if rownum == 1:
        continue
    data = (row[0].value, row[1].value, row[2].value, row[3].value, row[6].value, row[7].value )

    if row[4].value == "성공":
        continue

    print("email sending --> ", row[0].value, row[1].value)
    result = email_send(data)

    dt = datetime.datetime.now()
    if result:
        ws['E'+str(rownum)] = "성공"
        ws['F'+str(rownum)] = dt.strftime("%Y-%m-%d %H:%M:%S")
    else:
        ws['E'+str(rownum)] = "실패"
        ws['F'+str(rownum)] = dt.strftime("%Y-%m-%d %H:%M:%S")

wb.save("email_list.xlsx")